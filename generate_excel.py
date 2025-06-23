# generate_excel.py
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment

def generate_excel_with_charts(file_id, station_name, latitude, longitude, start_date, end_date, nasa_data, solar_data):
    ws2m = nasa_data["properties"]["parameter"]["WS2M"]
    wd2m = nasa_data["properties"]["parameter"]["WD2M"]

    df = pd.DataFrame({
        "Date": pd.to_datetime(list(ws2m.keys()), format="%Y%m%d"),
        "Wind Speed (m/s)": list(ws2m.values()),
        "Wind Direction (degrees)": list(wd2m.values())
    })

    df["YearMonth"] = df["Date"].dt.to_period("M")
    monthly_avg = df.groupby("YearMonth").agg({
        "Wind Speed (m/s)": "mean",
        "Wind Direction (degrees)": "mean"
    }).reset_index()
    monthly_avg["YearMonth"] = monthly_avg["YearMonth"].astype(str)

    degree_ticks = np.arange(0, 360, 10)

    def create_polar_chart(fig_id, theta, r, title, cmap, filename):
        fig = plt.figure(figsize=(6, 6))
        ax = fig.add_subplot(111, polar=True)
        scatter = ax.scatter(theta, r, c=r, cmap=cmap, alpha=0.75, edgecolors='black')
        plt.colorbar(scatter, ax=ax, label='Wind Speed (m/s)')
        ax.set_theta_zero_location('N')
        ax.set_theta_direction(-1)
        ax.set_xticks(np.deg2rad(degree_ticks))
        ax.set_xticklabels([f"{d}°" for d in degree_ticks])
        for label, degree in zip(['N', 'E', 'S', 'W'], [0, 90, 180, 270]):
            ax.text(np.deg2rad(degree), ax.get_rmax() * 1.1, label, ha='center', va='center', fontsize=12, fontweight='bold')
        ax.set_title(title, pad=20)
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()

    theta_daily = df["Wind Direction (degrees)"] * np.pi / 180.0
    r_daily = df["Wind Speed (m/s)"]

    scatter_chart = f"{file_id}_scatter.jpg"
    create_polar_chart(file_id, theta_daily, r_daily, f"Polar Wind Chart - {station_name}", 'viridis', scatter_chart)

    fig2 = plt.figure(figsize=(6, 6))
    ax2 = fig2.add_subplot(111, polar=True)
    sort_idx = np.argsort(theta_daily)
    ax2.bar(theta_daily.iloc[sort_idx], r_daily.iloc[sort_idx],
            width=np.deg2rad(8),
            bottom=0.0,
            color=plt.cm.viridis(r_daily.iloc[sort_idx] / r_daily.max()),
            alpha=0.75,
            edgecolor='black')
    ax2.set_theta_zero_location('N')
    ax2.set_theta_direction(-1)
    ax2.set_xticks(np.deg2rad(degree_ticks))
    ax2.set_xticklabels([f"{d}°" for d in degree_ticks])
    for label, degree in zip(['N', 'E', 'S', 'W'], [0, 90, 180, 270]):
        ax2.text(np.deg2rad(degree), ax2.get_rmax() * 1.1, label, ha='center', va='center', fontsize=12, fontweight='bold')
    ax2.set_title(f"Polar Wind Rose - {station_name}", pad=20)
    bar_chart = f"{file_id}_bar.jpg"
    plt.colorbar(plt.cm.ScalarMappable(cmap='viridis', norm=plt.Normalize(0, r_daily.max())), ax=ax2)
    plt.savefig(bar_chart, dpi=300, bbox_inches='tight')
    plt.close()

    theta_monthly = monthly_avg["Wind Direction (degrees)"] * np.pi / 180.0
    r_monthly = monthly_avg["Wind Speed (m/s)"]
    summary_chart = f"{file_id}_summary.jpg"
    create_polar_chart(file_id, theta_monthly, r_monthly, f"Monthly Summary Polar Chart - {station_name}", 'plasma', summary_chart)

    excel_filename = f"{file_id}_wind_data_with_charts.xlsx"
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df[["Date", "Wind Speed (m/s)", "Wind Direction (degrees)"]].to_excel(writer, sheet_name='Wind Data', index=False)
        monthly_avg.to_excel(writer, sheet_name='Monthly Summary', index=False)

        # Solar Radiation data
        solar_param = solar_data["properties"]["parameter"]["ALLSKY_SFC_SW_DWN"]
        df_solar = pd.DataFrame({
            "Date": pd.to_datetime(list(solar_param.keys()), format="%Y%m%d"),
            "Solar Radiation (kWh/m²/day)": list(solar_param.values())
        })
        df_solar.to_excel(writer, sheet_name='Solar Radiation', index=False)

        # Monthly Solar Radiation (Bar chart)
        df_solar["YearMonth"] = df_solar["Date"].dt.to_period("M")
        monthly_solar_avg = df_solar.groupby("YearMonth").agg({
            "Solar Radiation (kWh/m²/day)": "mean"
        }).reset_index()
        monthly_solar_avg["YearMonth"] = monthly_solar_avg["YearMonth"].astype(str)
        monthly_solar_avg.to_excel(writer, sheet_name='Monthly Solar Radiation', index=False)

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        info_df = pd.DataFrame({
            "Parameter": ["Station Name", "Latitude", "Longitude", "Start Date", "End Date", "Author", "Generated At", "Google Maps Link"],
            "Value": [
                station_name,
                latitude,
                longitude,
                start_date.strftime("%Y-%m-%d"),
                end_date.strftime("%Y-%m-%d"),
                "Sergio Cruz",
                now_str,
                f"https://www.google.com/maps?q={latitude},{longitude}"
            ]
        })
        info_df.to_excel(writer, sheet_name='Info', index=False)

    wb = load_workbook(excel_filename)

    def auto_width(sheet):
        for col in sheet.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = max_len + 2

    def format_cells(sheet, column_letter, number_format=None, align='center'):
        for cell in sheet[column_letter][1:]:
            if number_format:
                cell.number_format = number_format
            cell.alignment = Alignment(horizontal=align)

    auto_width(wb["Wind Data"])
    auto_width(wb["Monthly Summary"])
    auto_width(wb["Solar Radiation"])
    auto_width(wb["Monthly Solar Radiation"])
    auto_width(wb["Info"])

    ws = wb["Wind Data"]
    format_cells(ws, "A", "DD-MMM-YY", "center")
    format_cells(ws, "B", "0.00", "right")
    format_cells(ws, "C", "0", "center")

    ws2 = wb["Monthly Summary"]
    format_cells(ws2, "A", align="center")
    format_cells(ws2, "B", "0.00", "right")
    format_cells(ws2, "C", "0", "center")

    ws3 = wb["Solar Radiation"]
    format_cells(ws3, "A", "DD-MMM-YY", "center")
    format_cells(ws3, "B", "0.00", "right")

    ws4 = wb["Monthly Solar Radiation"]
    format_cells(ws4, "A", align="center")
    format_cells(ws4, "B", "0.00", "right")

    # Monthly Solar Radiation Chart
    fig_solar = plt.figure(figsize=(8, 6))
    ax_solar = fig_solar.add_subplot(111)
    ax_solar.barh(
        monthly_solar_avg["YearMonth"],
        monthly_solar_avg["Solar Radiation (kWh/m²/day)"],
        color='goldenrod',
        edgecolor='black'
    )
    ax_solar.set_xlabel("Solar Radiation (kWh/m²/day)")
    ax_solar.set_title(f"Monthly Solar Radiation - {station_name}")
    plt.tight_layout()

    solar_chart_img = f"{file_id}_solar_monthly.jpg"
    plt.savefig(solar_chart_img, dpi=300)
    plt.close()

    # Insert charts as sheets
    for title, img_file in [
        ("Scatter Chart", scatter_chart),
        ("Wind Rose Chart", bar_chart),
        ("Monthly Summary Polar", summary_chart),
        ("Monthly Solar Radiation Chart", solar_chart_img)
    ]:
        sheet = wb.create_sheet(title=title)
        img = ExcelImage(img_file)
        img.anchor = 'A1'
        sheet.add_image(img)

    # Reorder sheets as requested
    desired_order = [
        "Info",
        "Solar Radiation",
        "Monthly Solar Radiation",
        "Wind Data",
        "Monthly Summary",
        "Scatter Chart",
        "Wind Rose Chart",
        "Monthly Summary Polar",
        "Monthly Solar Radiation Chart"
    ]

    wb._sheets = [wb[sheet] for sheet in desired_order if sheet in wb.sheetnames]

    wb.save(excel_filename)

    os.remove(scatter_chart)
    os.remove(bar_chart)
    os.remove(summary_chart)
    os.remove(solar_chart_img)

    return excel_filename