from fastapi import FastAPI
from fastapi.responses import FileResponse
from pydantic import BaseModel
import matplotlib.pyplot as plt
import pandas as pd

import os
import uuid
import httpx

import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import numbers
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from datetime import date, datetime


# Load .env file
load_dotenv()

app = FastAPI()

ENVIRONMENT = os.getenv("ENVIRONMENT", "local")

if ENVIRONMENT == "production":
    allowed_origins = ["https://your-production-frontend.com"]
else:
    allowed_origins = ["http://localhost:3000", "null"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class WindDataRequest(BaseModel):
    station_name: str
    latitude: float
    longitude: float
    start: date
    end: date

@app.post("/generate-files")
async def generate_files(request: WindDataRequest):
    file_id = str(uuid.uuid4())

    nasa_api_url = (
        "https://power.larc.nasa.gov/api/temporal/daily/point"
        f"?parameters=WS2M,WD2M"
        f"&community=RE"
        f"&latitude={request.latitude}"
        f"&longitude={request.longitude}"
        f"&start={request.start.strftime('%Y%m%d')}"
        f"&end={request.end.strftime('%Y%m%d')}"
        "&format=JSON"
    )

    async with httpx.AsyncClient() as client:
        nasa_response = await client.get(nasa_api_url)

    if nasa_response.status_code != 200:
        return {"error": "Failed to fetch data from NASA POWER API"}

    nasa_json = nasa_response.json()
    daily_data = nasa_json["properties"]["parameter"]
    ws2m = daily_data["WS2M"]
    wd2m = daily_data["WD2M"]

    df_nasa = pd.DataFrame({
        "Date": pd.to_datetime(list(ws2m.keys()), format="%Y%m%d"),
        "Wind Speed (m/s)": list(ws2m.values()),
        "Wind Direction (degrees)": list(wd2m.values())
    })

    df_nasa["YearMonth"] = df_nasa["Date"].dt.to_period("M")

    monthly_avg = df_nasa.groupby("YearMonth").agg({
        "Wind Speed (m/s)": "mean",
        "Wind Direction (degrees)": "mean"
    }).reset_index()

    monthly_avg["YearMonth"] = monthly_avg["YearMonth"].astype(str)

    # Chart 1: Scatter Polar Chart (daily)
    fig1 = plt.figure(figsize=(6, 6))
    ax1 = fig1.add_subplot(111, polar=True)

    theta = df_nasa["Wind Direction (degrees)"] * (np.pi / 180.0)
    r = df_nasa["Wind Speed (m/s)"]

    scatter = ax1.scatter(theta, r, c=r, cmap='viridis', alpha=0.75)
    plt.colorbar(scatter, ax=ax1, label='Wind Speed (m/s)')

    ax1.set_theta_zero_location('N')
    ax1.set_theta_direction(-1)

    degree_ticks = np.arange(0, 360, 10)
    ax1.set_xticks(np.deg2rad(degree_ticks))
    ax1.set_xticklabels([f"{d}°" for d in degree_ticks])

    for label, degree in zip(['N', 'E', 'S', 'W'], [0, 90, 180, 270]):
        angle_rad = np.deg2rad(degree)
        ax1.text(angle_rad, ax1.get_rmax() + 0.1 * ax1.get_rmax(), label,
                 horizontalalignment='center', verticalalignment='center',
                 fontsize=12, fontweight='bold', color='black')

    ax1.set_title(f"Polar Wind Chart (Scatter) - {request.station_name}", pad=20)

    scatter_jpg_filename = f"{file_id}_wind_scatter_chart.jpg"
    plt.savefig(scatter_jpg_filename, dpi=300, bbox_inches='tight')
    plt.close()

    # Chart 2: Wind Rose Bar Chart
    fig2 = plt.figure(figsize=(6, 6))
    ax2 = fig2.add_subplot(111, polar=True)

    sort_idx = np.argsort(theta)
    theta_sorted = theta.iloc[sort_idx]
    r_sorted = r.iloc[sort_idx]

    bars = ax2.bar(theta_sorted, r_sorted,
                   width=np.deg2rad(8),
                   bottom=0.0,
                   color=plt.cm.viridis(r_sorted / r_sorted.max()),
                   alpha=0.75,
                   edgecolor='black')

    ax2.set_theta_zero_location('N')
    ax2.set_theta_direction(-1)

    ax2.set_xticks(np.deg2rad(degree_ticks))
    ax2.set_xticklabels([f"{d}°" for d in degree_ticks])

    for label, degree in zip(['N', 'E', 'S', 'W'], [0, 90, 180, 270]):
        angle_rad = np.deg2rad(degree)
        ax2.text(angle_rad, ax2.get_rmax() + 0.1 * ax2.get_rmax(), label,
                 horizontalalignment='center', verticalalignment='center',
                 fontsize=12, fontweight='bold', color='black')

    sm = plt.cm.ScalarMappable(cmap='viridis', norm=plt.Normalize(0, r_sorted.max()))
    sm.set_array([])
    plt.colorbar(sm, ax=ax2, pad=0.1, label='Wind Speed (m/s)')

    ax2.set_title(f"Polar Wind Rose (Bar Chart) - {request.station_name}", pad=20)

    bar_jpg_filename = f"{file_id}_wind_bar_chart.jpg"
    plt.savefig(bar_jpg_filename, dpi=300, bbox_inches='tight')
    plt.close()

    # Chart 3: Monthly Summary Polar Chart
    fig5 = plt.figure(figsize=(6,6))
    ax5 = fig5.add_subplot(111, polar=True)

    theta5 = monthly_avg["Wind Direction (degrees)"] * (np.pi / 180.0)
    r5 = monthly_avg["Wind Speed (m/s)"]

    scatter5 = ax5.scatter(theta5, r5, c=r5, cmap='plasma', s=100, edgecolors='black', alpha=0.8)

    for i, month in enumerate(monthly_avg["YearMonth"]):
        ax5.text(theta5.iloc[i], r5.iloc[i] + 0.1, month, fontsize=10, ha='center', va='center')

    plt.colorbar(scatter5, ax=ax5, pad=0.1, label='Wind Speed (m/s)')

    ax5.set_theta_zero_location('N')
    ax5.set_theta_direction(-1)

    ax5.set_xticks(np.deg2rad(degree_ticks))
    ax5.set_xticklabels([f"{d}°" for d in degree_ticks])

    for label, degree in zip(['N', 'E', 'S', 'W'], [0, 90, 180, 270]):
        angle_rad = np.deg2rad(degree)
        ax5.text(angle_rad, ax5.get_rmax() + 0.1 * ax5.get_rmax(), label,
                 horizontalalignment='center', verticalalignment='center',
                 fontsize=12, fontweight='bold', color='black')

    ax5.set_title(f"Monthly Summary Polar Chart - {request.station_name}", pad=20)

    monthly_summary_polar_chart = f"{file_id}_monthly_summary_polar.jpg"
    plt.savefig(monthly_summary_polar_chart, dpi=300, bbox_inches='tight')
    plt.close()

    # Generate Excel
    excel_filename = f"{file_id}_wind_data_with_charts.xlsx"

    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df_nasa[["Date", "Wind Speed (m/s)", "Wind Direction (degrees)"]].to_excel(writer, sheet_name='Wind Data', index=False)
        monthly_avg.to_excel(writer, sheet_name='Monthly Summary', index=False)
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        info_df = pd.DataFrame({
            "Parameter": ["Station Name", "Latitude", "Longitude", "Start Date", "End Date", "Author", "Generated At"],
            "Value": [
                request.station_name,
                request.latitude,
                request.longitude,
                request.start.strftime("%Y-%m-%d"),
                request.end.strftime("%Y-%m-%d"),
                "Sergio Cruz",
                now_str
            ]
        })
        info_df.to_excel(writer, sheet_name='Info', index=False)

    wb = load_workbook(excel_filename)

    # Adjust columns and format Date column
    ws_wind = wb["Wind Data"]
    for column_cells in ws_wind.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws_wind.column_dimensions[column_cells[0].column_letter].width = length + 2

    for cell in ws_wind["A"][1:]:
        cell.number_format = 'DD-MMM-YY'

    for sheet_name in ["Monthly Summary", "Info"]:
        ws = wb[sheet_name]
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Add charts
    ws_scatter = wb.create_sheet(title='Scatter Chart')
    img1 = ExcelImage(scatter_jpg_filename)
    img1.anchor = 'A1'
    ws_scatter.add_image(img1)

    ws_bar = wb.create_sheet(title='Wind Rose Chart')
    img2 = ExcelImage(bar_jpg_filename)
    img2.anchor = 'A1'
    ws_bar.add_image(img2)

    ws_summary_polar = wb.create_sheet(title='Monthly Summary Polar')
    img5 = ExcelImage(monthly_summary_polar_chart)
    img5.anchor = 'A1'
    ws_summary_polar.add_image(img5)

    wb.save(excel_filename)

    os.remove(scatter_jpg_filename)
    os.remove(bar_jpg_filename)
    os.remove(monthly_summary_polar_chart)

    return {
        "excel_file_url": f"/download/{excel_filename}"
    }

@app.get("/download/{filename}")
async def download_file(filename: str):
    file_path = os.path.join(".", filename)
    if os.path.isfile(file_path):
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    return {"error": "File not found"}
